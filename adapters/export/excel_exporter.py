"""
Exportador de actividades a Excel.
Implementa la lógica de exportación separada de la UI.
"""

import subprocess
import sys
from pathlib import Path
from typing import Any


def _ensure_openpyxl() -> bool:
    """Verifica/openpyxl. Instala si no está. Retorna True si está disponible."""
    try:
        from openpyxl import Workbook  # noqa: F401
        return True
    except ImportError:
        try:
            subprocess.run(
                [sys.executable, "-m", "pip", "install", "openpyxl"],
                check=True,
                capture_output=True,
                timeout=60,
            )
            return True
        except Exception:
            return False


class ExcelActivityExporter:
    """Exporta los 3 reportes (Tareas, Subtareas, Transiciones) a Excel."""

    def export(
        self,
        activities: list[dict[str, Any]],
        subtasks: list[dict[str, Any]],
        transitions: list[dict[str, Any]],
        filepath: Path,
    ) -> bool:
        """
        Exporta Tareas, Subtareas y Transiciones a un archivo Excel con 3 hojas.
        Retorna True si tuvo éxito.
        """
        if not _ensure_openpyxl():
            return False

        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws_tareas = wb.active
        ws_tareas.title = "Tareas"
        headers_tareas = [
            "ID", "Ticket", "Actividad", "Estado", "Subtareas (hechas/total)",
            "Inicio columna", "Inicio In Progress", "Fin (Done)"
        ]
        for col, h in enumerate(headers_tareas, 1):
            ws_tareas.cell(row=1, column=col, value=h).font = Font(bold=True)
        for row_idx, t in enumerate(activities, 2):
            subtasks_list = t.get("subtasks", [])
            done = sum(1 for s in subtasks_list if s.get("done"))
            sub_str = f"{done}/{len(subtasks_list)}" if subtasks_list else "-"
            ws_tareas.cell(row=row_idx, column=1, value=t.get("id", ""))
            ws_tareas.cell(row=row_idx, column=2, value=t.get("ticket", ""))
            ws_tareas.cell(row=row_idx, column=3, value=t.get("name", ""))
            ws_tareas.cell(row=row_idx, column=4, value=t.get("column_display", ""))
            ws_tareas.cell(row=row_idx, column=5, value=sub_str)
            entered = t.get("entered_at", "")[:19] if t.get("entered_at") else ""
            ws_tareas.cell(row=row_idx, column=6, value=entered)
            started = t.get("started_at", "")[:19] if t.get("started_at") else "-"
            ws_tareas.cell(row=row_idx, column=7, value=started)
            finished = t.get("finished_at", "")[:19] if t.get("finished_at") else "-"
            ws_tareas.cell(row=row_idx, column=8, value=finished)
        for col in range(1, 9):
            ws_tareas.column_dimensions[get_column_letter(col)].width = 18
        ws_tareas.column_dimensions["C"].width = 40

        # Hoja 2: Subtareas
        ws_sub = wb.create_sheet("Subtareas", 1)
        headers_sub = ["Tarea ID", "Ticket", "Tarea", "Subtarea", "Hecha", "Estado tarea"]
        for col, h in enumerate(headers_sub, 1):
            ws_sub.cell(row=1, column=col, value=h).font = Font(bold=True)
        for row_idx, s in enumerate(subtasks, 2):
            ws_sub.cell(row=row_idx, column=1, value=s.get("task_id", ""))
            ws_sub.cell(row=row_idx, column=2, value=s.get("task_ticket", ""))
            ws_sub.cell(row=row_idx, column=3, value=s.get("task_name", ""))
            ws_sub.cell(row=row_idx, column=4, value=s.get("subtask_text", ""))
            ws_sub.cell(row=row_idx, column=5, value="Sí" if s.get("done") else "No")
            ws_sub.cell(row=row_idx, column=6, value=s.get("column_display", ""))
        for col in range(1, 7):
            ws_sub.column_dimensions[get_column_letter(col)].width = 20
        ws_sub.column_dimensions["D"].width = 45

        # Hoja 3: Transiciones
        ws_trans = wb.create_sheet("Transiciones", 2)
        headers_trans = ["Tarea ID", "Tarea", "Desde", "Hacia", "Fecha/hora"]
        for col, h in enumerate(headers_trans, 1):
            ws_trans.cell(row=1, column=col, value=h).font = Font(bold=True)
        for row_idx, tr in enumerate(transitions, 2):
            ws_trans.cell(row=row_idx, column=1, value=tr.get("task_id", ""))
            ws_trans.cell(row=row_idx, column=2, value=tr.get("task_name", ""))
            ws_trans.cell(row=row_idx, column=3, value=tr.get("from_display", ""))
            ws_trans.cell(row=row_idx, column=4, value=tr.get("to_display", ""))
            ts = tr.get("timestamp", "")
            ws_trans.cell(row=row_idx, column=5, value=ts[:19] if ts else "")
        for col in range(1, 6):
            ws_trans.column_dimensions[get_column_letter(col)].width = 22
        ws_trans.column_dimensions["B"].width = 40

        wb.save(filepath)
        return True
