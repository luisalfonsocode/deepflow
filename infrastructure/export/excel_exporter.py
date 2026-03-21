"""
Exportador de actividades a Excel.
Implementa la lógica de exportación separada de la UI.
Requisito: openpyxl en requirements.txt.
"""

import logging
from pathlib import Path
from typing import Any

LOG = logging.getLogger(__name__)

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class ExcelActivityExporter:
    """Exporta los 3 reportes (Tareas, Subtareas, Transiciones) a Excel."""

    def export(
        self,
        activities: list[dict[str, Any]],
        subtasks: list[dict[str, Any]],
        transitions: list[dict[str, Any]],
        filepath: Path,
    ) -> bool:
        """
        Exporta Tareas, Subtareas y Transiciones a un archivo Excel con 3 hojas.
        Retorna True si tuvo éxito.
        """
        LOG.info("Exportando a Excel: %s (tareas=%d, subtareas=%d, transiciones=%d)",
                 filepath, len(activities), len(subtasks), len(transitions))
        wb = Workbook()
        ws_tareas = wb.active
        ws_tareas.title = "Tareas"
        headers_tareas = [
            "ID", "Ticket", "Actividad", "Estado", "Tribe & Squad", "Requester",
            "Reporting channel", "Subtareas (hechas/total)",
            "Fecha solicitud", "Fecha inicio", "Fecha fin", "Fecha compromiso"
        ]
        for col, h in enumerate(headers_tareas, 1):
            ws_tareas.cell(row=1, column=col, value=h).font = Font(bold=True)
        for row_idx, t in enumerate(activities, 2):
            subtasks_list = t.get("subtasks", [])
            done = sum(1 for s in subtasks_list if s.get("done"))
            sub_str = f"{done}/{len(subtasks_list)}" if subtasks_list else "-"
            ws_tareas.cell(row=row_idx, column=1, value=t.get("id", ""))
            ws_tareas.cell(row=row_idx, column=2, value=t.get("ticket", ""))
            ws_tareas.cell(row=row_idx, column=3, value=t.get("name", ""))
            ws_tareas.cell(row=row_idx, column=4, value=t.get("column_display", ""))
            ws_tareas.cell(row=row_idx, column=5, value=t.get("tribe_and_squad", ""))
            ws_tareas.cell(row=row_idx, column=6, value=t.get("requester", ""))
            ws_tareas.cell(row=row_idx, column=7, value=t.get("reporting_channel", ""))
            ws_tareas.cell(row=row_idx, column=8, value=sub_str)
            created = (t.get("created_at") or t.get("entered_at") or "")[:19]
            ws_tareas.cell(row=row_idx, column=9, value=created or "-")
            started = t.get("started_at", "")[:19] if t.get("started_at") else "-"
            ws_tareas.cell(row=row_idx, column=10, value=started)
            finished = t.get("finished_at", "")[:19] if t.get("finished_at") else "-"
            ws_tareas.cell(row=row_idx, column=11, value=finished)
            due = t.get("due_date", "")[:19] if t.get("due_date") else "-"
            ws_tareas.cell(row=row_idx, column=12, value=due)
        for col in range(1, 13):
            ws_tareas.column_dimensions[get_column_letter(col)].width = 18
        ws_tareas.column_dimensions["C"].width = 40
        ws_tareas.column_dimensions["E"].width = 20
        ws_tareas.column_dimensions["F"].width = 20
        ws_tareas.column_dimensions["G"].width = 18

        # Hoja 2: Subtareas
        ws_sub = wb.create_sheet("Subtareas", 1)
        headers_sub = ["Tarea ID", "Ticket", "Tarea", "Subtarea", "Hecha", "Estado tarea"]
        for col, h in enumerate(headers_sub, 1):
            ws_sub.cell(row=1, column=col, value=h).font = Font(bold=True)
        for row_idx, s in enumerate(subtasks, 2):
            ws_sub.cell(row=row_idx, column=1, value=s.get("task_id", ""))
            ws_sub.cell(row=row_idx, column=2, value=s.get("task_ticket", ""))
            ws_sub.cell(row=row_idx, column=3, value=s.get("task_name", ""))
            ws_sub.cell(row=row_idx, column=4, value=s.get("subtask_text", ""))
            ws_sub.cell(row=row_idx, column=5, value="Sí" if s.get("done") else "No")
            ws_sub.cell(row=row_idx, column=6, value=s.get("column_display", ""))
        for col in range(1, 7):
            ws_sub.column_dimensions[get_column_letter(col)].width = 20
        ws_sub.column_dimensions["D"].width = 45

        # Hoja 3: Transiciones
        ws_trans = wb.create_sheet("Transiciones", 2)
        headers_trans = ["Tarea ID", "Tarea", "Desde", "Hacia", "Fecha/hora"]
        for col, h in enumerate(headers_trans, 1):
            ws_trans.cell(row=1, column=col, value=h).font = Font(bold=True)
        for row_idx, tr in enumerate(transitions, 2):
            ws_trans.cell(row=row_idx, column=1, value=tr.get("task_id", ""))
            ws_trans.cell(row=row_idx, column=2, value=tr.get("task_name", ""))
            ws_trans.cell(row=row_idx, column=3, value=tr.get("from_display", ""))
            ws_trans.cell(row=row_idx, column=4, value=tr.get("to_display", ""))
            ts = tr.get("timestamp", "")
            ws_trans.cell(row=row_idx, column=5, value=ts[:19] if ts else "")
        for col in range(1, 6):
            ws_trans.column_dimensions[get_column_letter(col)].width = 22
        ws_trans.column_dimensions["B"].width = 40

        wb.save(filepath)
        LOG.info("Exportación Excel completada: %s", filepath)
        return True

    def export_with_time_report(
        self,
        activities: list[dict[str, Any]],
        subtasks: list[dict[str, Any]],
        transitions: list[dict[str, Any]],
        time_report: dict[str, Any],
        filepath: Path,
    ) -> bool:
        """
        Exporta Tareas, Subtareas, Transiciones y Reporte de tiempo (hoja extra).
        El reporte incluye: resumen por categoría (activo/detenido/%) y detalle por tarea.
        """
        if not self.export(activities, subtasks, transitions, filepath):
            return False
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        ws_time = wb.create_sheet("Reporte de tiempo", 3)
        from_date = time_report.get("from_date", "")
        to_date = time_report.get("to_date", "")
        if hasattr(from_date, "strftime"):
            from_date = from_date.strftime("%Y-%m-%d")
        if hasattr(to_date, "strftime"):
            to_date = to_date.strftime("%Y-%m-%d")
        ws_time.cell(row=1, column=1, value="Periodo:").font = Font(bold=True)
        ws_time.cell(row=1, column=2, value=f"{from_date} a {to_date}")
        ws_time.cell(row=2, column=1, value="")
        row = 3
        headers_sum = ["Categoría", "Tiempo activo", "Tiempo detenido", "% Total", "Tareas", "Tiempo (días)"]
        for col, h in enumerate(headers_sum, 1):
            ws_time.cell(row=row, column=col, value=h).font = Font(bold=True)
        row += 1
        for s in time_report.get("summary_by_categoria", []):
            ws_time.cell(row=row, column=1, value=s.get("categoria", ""))
            ws_time.cell(row=row, column=2, value=s.get("active_fmt", ""))
            ws_time.cell(row=row, column=3, value=s.get("detenido_fmt", ""))
            ws_time.cell(row=row, column=4, value=f"{s.get('pct_total', 0)}%")
            ws_time.cell(row=row, column=5, value=s.get("task_count", 0))
            ws_time.cell(row=row, column=6, value=s.get("tiempo_dias", 0))
            row += 1
        row += 2
        headers_det = ["Tarea", "Ticket", "Categoría", "Tiempo activo", "Tiempo detenido"]
        for col, h in enumerate(headers_det, 1):
            ws_time.cell(row=row, column=col, value=h).font = Font(bold=True)
        row += 1
        for d in time_report.get("detail_by_task", []):
            ws_time.cell(row=row, column=1, value=(d.get("name", ""))[:80])
            ws_time.cell(row=row, column=2, value=d.get("ticket", ""))
            ws_time.cell(row=row, column=3, value=d.get("categoria", ""))
            ws_time.cell(row=row, column=4, value=d.get("active_fmt", ""))
            ws_time.cell(row=row, column=5, value=d.get("detenido_fmt", ""))
            row += 1
        for c in range(1, 7):
            ws_time.column_dimensions[get_column_letter(c)].width = 22
        ws_time.column_dimensions["A"].width = 45
        wb.save(filepath)
        LOG.info("Exportación con reporte de tiempo: %s", filepath)
        return True

    def export_time_summary_only(self, time_report: dict[str, Any], filepath: Path) -> bool:
        """Exporta solo resumen por categoría a Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen por categoría"
        from_date = time_report.get("from_date", "")
        to_date = time_report.get("to_date", "")
        if hasattr(from_date, "strftime"):
            from_date = from_date.strftime("%Y-%m-%d")
        if hasattr(to_date, "strftime"):
            to_date = to_date.strftime("%Y-%m-%d")
        ws.cell(row=1, column=1, value="Periodo:").font = Font(bold=True)
        ws.cell(row=1, column=2, value=f"{from_date} a {to_date}")
        row = 3
        headers = ["Categoría", "Tiempo activo", "Tiempo detenido", "% Total", "Tareas", "Tiempo (días)"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h).font = Font(bold=True)
        row += 1
        for s in time_report.get("summary_by_categoria", []):
            ws.cell(row=row, column=1, value=s.get("categoria", ""))
            ws.cell(row=row, column=2, value=s.get("active_fmt", ""))
            ws.cell(row=row, column=3, value=s.get("detenido_fmt", ""))
            ws.cell(row=row, column=4, value=f"{s.get('pct_total', 0)}%")
            ws.cell(row=row, column=5, value=s.get("task_count", 0))
            ws.cell(row=row, column=6, value=s.get("tiempo_dias", 0))
            row += 1
        for c in range(1, 7):
            ws.column_dimensions[get_column_letter(c)].width = 22
        ws.column_dimensions["A"].width = 45
        wb.save(filepath)
        LOG.info("Exportación resumen tiempo: %s", filepath)
        return True

    def export_time_detail_only(self, time_report: dict[str, Any], filepath: Path) -> bool:
        """Exporta solo detalle por tarea a Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Detalle por tarea"
        from_date = time_report.get("from_date", "")
        to_date = time_report.get("to_date", "")
        if hasattr(from_date, "strftime"):
            from_date = from_date.strftime("%Y-%m-%d")
        if hasattr(to_date, "strftime"):
            to_date = to_date.strftime("%Y-%m-%d")
        ws.cell(row=1, column=1, value="Periodo:").font = Font(bold=True)
        ws.cell(row=1, column=2, value=f"{from_date} a {to_date}")
        row = 3
        headers = ["Tarea", "Ticket", "Categoría", "Tiempo activo", "Tiempo detenido"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h).font = Font(bold=True)
        row += 1
        for d in time_report.get("detail_by_task", []):
            ws.cell(row=row, column=1, value=(d.get("name", ""))[:80])
            ws.cell(row=row, column=2, value=d.get("ticket", ""))
            ws.cell(row=row, column=3, value=d.get("categoria", ""))
            ws.cell(row=row, column=4, value=d.get("active_fmt", ""))
            ws.cell(row=row, column=5, value=d.get("detenido_fmt", ""))
            row += 1
        for c in range(1, 6):
            ws.column_dimensions[get_column_letter(c)].width = 22
        ws.column_dimensions["A"].width = 45
        wb.save(filepath)
        LOG.info("Exportación detalle tiempo: %s", filepath)
        return True
